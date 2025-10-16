import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta
import os
import sys
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def extract_date(value):
    if isinstance(value, str):
        match = re.search(r"\d{1,2}\.\d{1,2}\.\d{4}", value)
        if match:
            return match.group(0)
    return None

def split_row(row):
    try:
        start = datetime.combine(row["Datum"].date(), row["Startzeit"])
        end = datetime.combine(row["Datum"].date(), row["Endzeit"])
        # Fix für 00:00 → 00:00 mit Dauer > 0
        if start == end and row["Dauer netto (dezimal)"] > 0:
            end = start + timedelta(hours=row["Dauer netto (dezimal)"])
        elif end <= start:
            end += timedelta(days=1)
    except Exception:
        return []

    result = []
    current = start.replace(minute=0, second=0, microsecond=0)
    if current > start:
        current -= timedelta(hours=1)

    while current < end:
        next_hour = current + timedelta(hours=1)
        overlap_start = max(current, start)
        overlap_end = min(next_hour, end)
        minutes_in_hour = (overlap_end - overlap_start).total_seconds() / 60

        if minutes_in_hour > 0:
            ratio = minutes_in_hour / 60
            result.append({
                "Datum": start.date(),
                "Stunde": current.strftime("%H:00"),
                "Personalstunden": ratio
            })
        current = next_hour

    return result

def calculate_hourly_presence(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
    except Exception as e:
        messagebox.showerror("Fehler", f"Datei konnte nicht geöffnet werden:\n{str(e)}")
        return

    expected_sheet = "Alle Mitarbeiter"
    if expected_sheet not in sheet_names:
        messagebox.showerror("Fehlendes Blatt", f"'{expected_sheet}' nicht gefunden.")
        return

    df = pd.read_excel(file_path, sheet_name=expected_sheet, skiprows=6)
    df.columns = df.columns.str.strip()

    if "Typ" not in df.columns:
        messagebox.showerror("Fehler", "Spalte 'Typ' fehlt im Export.")
        return

    df = df[df["Typ"] == "A"]

    expected_cols = ["Tag", "Startzeit", "Endzeit", "Dauer netto (dezimal)"]
    df = df[[col for col in expected_cols if col in df.columns]]
    df = df[df["Tag"].notna()]
    df["Datum"] = df["Tag"].apply(extract_date)
    df = df[df["Datum"].notna()]
    df["Datum"] = pd.to_datetime(df["Datum"], dayfirst=True, errors="coerce")
    df = df[df["Datum"].notna()]

    df["Startzeit"] = pd.to_datetime(df["Startzeit"], format="%H:%M", errors="coerce").dt.time
    df["Endzeit"] = pd.to_datetime(df["Endzeit"], format="%H:%M", errors="coerce").dt.time
    df = df[df["Startzeit"].notna() & df["Endzeit"].notna()]
    df["Dauer netto (dezimal)"] = pd.to_numeric(df["Dauer netto (dezimal)"], errors="coerce")
    df = df[df["Dauer netto (dezimal)"].notna()]

    rows = []
    for _, row in df.iterrows():
        result = split_row(row)
        if result:
            rows.extend(result)

    if not rows:
        messagebox.showerror("Keine verwertbaren Daten", "Keine gültigen Arbeitszeiten gefunden.")
        return

    df_expanded = pd.DataFrame(rows)
    df_grouped = df_expanded.groupby(["Datum", "Stunde"]).sum(numeric_only=True).reset_index()
    df_grouped["Personalstunden"] = df_grouped["Personalstunden"].round(2)

    def sort_key(hour_str):
        hour = int(hour_str.split(":")[0])
        return hour + (24 if hour < 5 else 0)

    df_grouped["sort_index"] = df_grouped["Stunde"].apply(sort_key)
    df_grouped = df_grouped.sort_values(["Datum", "sort_index"]).drop(columns=["sort_index"])

    # Export vorbereiten
    output_rows = []
    for datum in df_grouped["Datum"].unique():
        output_rows.append([""])
        output_rows.append([datum.strftime("%Y-%m-%d")])
        for _, row in df_grouped[df_grouped["Datum"] == datum].iterrows():
            output_rows.append(["", row["Stunde"], row["Personalstunden"]])

    result_df = pd.DataFrame(output_rows)
    original_name = os.path.basename(file_path)
    base_name = os.path.splitext(original_name)[0]
    export_path = os.path.join(os.path.dirname(file_path), f"Stundenanalyse_{base_name}.xlsx")

    result_df.to_excel(export_path, index=False, header=["Datum", "Stunde", "Personalstunden"])

    wb = load_workbook(export_path)
    ws = wb.active
    for col in ["B", "C"]:
        ws.column_dimensions[col].width = 20
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    wb.save(export_path)

    messagebox.showinfo("Fertig", f"Auswertung gespeichert:\n{export_path}")
    sys.exit()

def main():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Bitte Export auswählen",
        filetypes=[("Excel-Dateien", "*.xlsx *.xls")]
    )
    if file_path:
        try:
            calculate_hourly_presence(file_path)
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler bei der Verarbeitung:\n{str(e)}")

if __name__ == "__main__":
    main()
